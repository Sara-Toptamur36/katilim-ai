"""Altin Veri Seti'ni Excel'den JSON'a cevirir.

NEDEN GEREKLI: Excel insanlarin doldurmasi icin kolay, ama testler ve
Extraction Accuracy hesabi JSON okur. Bu betik ikisi arasindaki koprudur.

KULLANIM:
    python gold_dataset/excel_to_json.py

Excel'i her guncelledikten sonra bunu calistir ve uretilen JSON'u commit'le.
CI, regresyon testlerinde bu JSON'u kullanir.

---------------------------------------------------------------------------
BOS HUCRE NE ZAMAN "KAYNAKTA YOK" SAYILIR (olcumun dogrulugu buna bagli)
---------------------------------------------------------------------------
Excel'in "1. Nasil Doldurulur" sayfasindaki kural nettir ve GECERLIDIR:
bilgi sayfada yoksa hucre BOS birakilir; '-', 'yok', '0' YAZILMAZ.

Ama bu kural yalnizca etiketleyicinin GERCEKTEN BAKTIGI sutunlar icin
anlamlidir. Sonradan eklenen bir sutunda tum hucreler dogal olarak bostur
ve bu "kaynakta yok" DEMEK DEGILDIR - sadece "henuz kimse bakmadi"
demektir. Ikisini karistirmak, hic etiketlenmemis bir sutunu "hepsi bos,
demek ki motor hic uydurmuyor" diye BEDAVA yuksek puana cevirirdi.

Bu yuzden bos hucrenin anlamini SUTUN belirler:

  * INCELENMIS_ALANLAR : bir etiketleme oturumunda tek tek gozden
    gecirilmis sutunlar. Bos hucre = "kaynakta belirtilmemis" ->
    `alan_belirtilmemis` bayragi konur, yanlis pozitif OLCULEBILIR.

  * Listede olmayan sutunlar : olcum disi. scraper/scripts/
    extraction_accuracy.py bunlarda yanlis pozitif saymaz.

BIR SUTUNUN ETIKETLEMESI BITINCE yapilacak tek is: sutun adini
INCELENMIS_ALANLAR'a eklemek. Etiketlemeyi hizlandirmak icin:
    python gold_dataset/etiketleme_yardimcisi.py
"""

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

KLASOR = Path(__file__).parent
EXCEL = KLASOR / "altin_veri_seti.xlsx"
JSON_CIKTI = KLASOR / "altin_veri_seti.json"

SAYFA = "2. Altin Veri Seti"
VERI_BASLANGIC_SATIRI = 3  # 1: baslik, 2: aciklama

SAYISAL_ALANLAR = {
    "kar_payi_orani",
    "maliyet_orani",
    "vade_ay",
    "finansman_tutari",
    "odul_miktari",
    # Vade ile KARISTIRILMAMALI - ucu de ayri kavram (bkz.
    # extraction/regex_extractor.py: "12 aya varan taksit" vade DEGILDIR).
    "taksit_sayisi",
    "erteleme_suresi_ay",
}
TARIH_ALANLARI = {"kampanya_baslangic", "kampanya_bitis", "giris_tarihi"}

GECERLI_PERIYOTLAR = {"aylik", "yillik", "belirsiz"}

# Etiketleme oturumunda tek tek gozden gecirilmis sutunlar: burada bos
# hucre "kaynakta belirtilmemis" demektir (bkz. modul docstring'i).
# 28 Temmuz 2026 oturumunun kapsami.
INCELENMIS_ALANLAR = (
    "kar_payi_orani",
    "vade_ay",
    "odul_miktari",
    "masraf_durumu",
    "kampanya_bitis",
    "hedef_kitle",
)

# Semada/Excel'de VAR ama henuz bir etiketleme oturumundan gecmemis
# sutunlar. Bos hucreleri "kaynakta yok" SAYILMAZ - olcum disidir.
# Etiketlemesi biten sutun buradan cikarilip INCELENMIS_ALANLAR'a eklenir.
#
# taksit_sayisi / erteleme_suresi_ay : sutunlar 9 Agustos'ta eklendi.
# odul_birimi / finansman_tutari     : 28 Temmuz oturumunda doldurulmus
#   ama bos birakilanlar gozden gecirilmemis (30 ve 16 kayitta deger var,
#   kalanlar icin "bakildi mi" bilinmiyor).
INCELENMEMIS_ALANLAR = (
    "taksit_sayisi",
    "erteleme_suresi_ay",
    "odul_birimi",
    "finansman_tutari",
)


class DogrulamaHatasi(Exception):
    """Altin Veri Seti'nde duzeltilmesi gereken bir sorun bulundu."""


def _tarihe_cevir(deger, kayit_id: str, alan: str) -> str | None:
    if deger is None or str(deger).strip() == "":
        return None
    if isinstance(deger, (datetime, date)):
        return deger.strftime("%Y-%m-%d")
    metin = str(deger).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", metin):
        return metin
    raise DogrulamaHatasi(
        f"[{kayit_id}] {alan}: tarih formati YYYY-AA-GG olmali, bulunan: {metin!r}"
    )


def _sayiya_cevir(deger, kayit_id: str, alan: str) -> float | int | None:
    if deger is None or str(deger).strip() == "":
        return None
    if isinstance(deger, (int, float)):
        return deger

    metin = str(deger).strip()
    # Sik yapilan hatalari yakala ve ACIK hata ver (sessizce duzeltme)
    if "%" in metin or "TL" in metin.upper() or re.search(r"[A-Za-zıİşŞğĞüÜöÖçÇ]", metin):
        raise DogrulamaHatasi(
            f"[{kayit_id}] {alan}: yalnizca sayi yazilmali (birim/isaret olmadan). "
            f"Bulunan: {metin!r}. Ornek dogru yazim: 1.89"
        )
    if "," in metin:
        raise DogrulamaHatasi(
            f"[{kayit_id}] {alan}: ondalik ayraci NOKTA olmali. "
            f"Bulunan: {metin!r} -> {metin.replace(',', '.')!r} olmali"
        )
    try:
        return float(metin) if "." in metin else int(metin)
    except ValueError:
        raise DogrulamaHatasi(f"[{kayit_id}] {alan}: sayiya cevrilemedi: {metin!r}")


def _kaydi_dogrula(kayit: dict) -> list[str]:
    """Mantiksal tutarlilik kontrolleri. Hata degil, UYARI listesi doner."""
    uyarilar = []
    kid = kayit.get("kayit_id", "?")

    oran = kayit.get("kar_payi_orani")
    periyot = kayit.get("oran_periyodu")

    if oran is not None and not periyot:
        uyarilar.append(
            f"[{kid}] Oran girilmis ama oran_periyodu bos. "
            "Aylik mi yillik mi belirtilmeli (bilinmiyorsa 'belirsiz')."
        )

    if periyot and periyot not in GECERLI_PERIYOTLAR:
        uyarilar.append(
            f"[{kid}] oran_periyodu gecersiz: {periyot!r}. "
            f"Gecerli: {', '.join(sorted(GECERLI_PERIYOTLAR))}"
        )

    # Aylik oran cok yuksekse muhtemelen yillik oran aylik sanilmis
    if oran is not None and periyot == "aylik" and oran > 20:
        uyarilar.append(
            f"[{kid}] Aylik oran %{oran} olamaz - bu muhtemelen YILLIK bir orandir "
            "(ya da mevduat orani). Kaynak sayfayi tekrar kontrol et."
        )

    if kayit.get("odul_miktari") is not None and not kayit.get("odul_birimi"):
        uyarilar.append(f"[{kid}] odul_miktari var ama odul_birimi bos (Mil/Gram/TL...)")

    bas, bit = kayit.get("kampanya_baslangic"), kayit.get("kampanya_bitis")
    if bas and bit and bas > bit:
        uyarilar.append(f"[{kid}] kampanya_baslangic ({bas}) bitisten ({bit}) sonra")

    if not kayit.get("kaynak_url"):
        uyarilar.append(f"[{kid}] kaynak_url bos - provenance icin zorunlu")

    return uyarilar


def donustur(excel_yolu: Path = EXCEL) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(excel_yolu, data_only=True)
    if SAYFA not in wb.sheetnames:
        raise DogrulamaHatasi(f"'{SAYFA}' sayfasi bulunamadi. Sayfalar: {wb.sheetnames}")

    ws = wb[SAYFA]
    basliklar = [h.value for h in ws[1]]

    kayitlar: list[dict] = []
    tum_uyarilar: list[str] = []
    gorulen_idler: set[str] = set()

    for satir in ws.iter_rows(min_row=VERI_BASLANGIC_SATIRI, values_only=True):
        ham = dict(zip(basliklar, satir))

        kayit_id = (ham.get("kayit_id") or "").strip() if ham.get("kayit_id") else ""
        if not kayit_id:
            continue  # bos satir

        if kayit_id in gorulen_idler:
            tum_uyarilar.append(f"[{kayit_id}] Ayni kayit_id birden fazla satirda")
        gorulen_idler.add(kayit_id)

        kayit: dict = {}
        for alan, deger in ham.items():
            if alan is None:
                continue
            if alan in SAYISAL_ALANLAR:
                kayit[alan] = _sayiya_cevir(deger, kayit_id, alan)
            elif alan in TARIH_ALANLARI:
                kayit[alan] = _tarihe_cevir(deger, kayit_id, alan)
            else:
                kayit[alan] = str(deger).strip() if deger is not None else None

        # Seffaflik bayragi: YALNIZCA incelenmis sutunlarda bos hucre
        # "kaynakta belirtilmemis" sayilir (bkz. modul docstring'i).
        kayit["alan_belirtilmemis"] = {
            a: True for a in INCELENMIS_ALANLAR if kayit.get(a) is None
        }

        tum_uyarilar.extend(_kaydi_dogrula(kayit))
        kayitlar.append(kayit)

    return kayitlar, tum_uyarilar


def main() -> int:
    if not EXCEL.exists():
        print(f"HATA: {EXCEL} bulunamadi")
        return 1

    try:
        kayitlar, uyarilar = donustur()
    except DogrulamaHatasi as e:
        print(f"\nDUZELTILMESI GEREKEN HATA:\n  {e}\n")
        return 1

    ornek = [k for k in kayitlar if k.get("giren_kisi") == "ORNEK"]
    gercek = [k for k in kayitlar if k.get("giren_kisi") != "ORNEK"]

    with open(JSON_CIKTI, "w", encoding="utf-8") as f:
        json.dump(kayitlar, f, ensure_ascii=False, indent=2)

    print(f"Yazildi: {JSON_CIKTI}")
    print(f"  Toplam kayit : {len(kayitlar)}")
    print(f"  Ornek (ORNEK): {len(ornek)}")
    print(f"  Gercek veri  : {len(gercek)}")

    bankalar: dict[str, int] = {}
    for k in gercek:
        bankalar[k.get("banka") or "?"] = bankalar.get(k.get("banka") or "?", 0) + 1
    if bankalar:
        print("\n  Banka basina gercek kayit:")
        for b, n in sorted(bankalar.items()):
            durum = "OK" if n >= 5 else f"yetersiz (hedef 5-8)"
            print(f"    {b:24} {n:2}  {durum}")

    # Olcum kapsami: yanlis pozitif YALNIZCA incelenmis sutunlarda
    # olculebilir (bkz. modul docstring'i).
    print("\n  Yanlis pozitif olcum kapsami:")
    for alan in INCELENMIS_ALANLAR:
        bayrakli = sum(1 for k in gercek if k["alan_belirtilmemis"].get(alan) is True)
        print(f"    {alan:20} olculebilir={bayrakli:2}/{len(gercek)}  (incelendi)")
    for alan in INCELENMEMIS_ALANLAR:
        dolu = sum(1 for k in gercek if k.get(alan) is not None)
        print(
            f"    {alan:20} olculebilir= 0/{len(gercek)}  "
            f"(OLCUM DISI - {dolu} kayitta deger var, bos kalanlar incelenmedi)"
        )
    if INCELENMEMIS_ALANLAR:
        print(
            "\n    Bir sutunun etiketlemesi bitince adini excel_to_json.py'deki\n"
            "    INCELENMIS_ALANLAR listesine tasi - yanlis pozitif o an olculur\n"
            "    hale gelir. Yardimci: python gold_dataset/etiketleme_yardimcisi.py"
        )

    if uyarilar:
        print(f"\n  UYARILAR ({len(uyarilar)}):")
        for u in uyarilar:
            print(f"    - {u}")
    else:
        print("\n  Uyari yok.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
